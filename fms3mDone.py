import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

st.set_page_config(layout="wide", page_title="Dynamic 3-Statement Modeler")

st.title("📊 Dynamic 3-Statement Financial Modeler")

# Define quarters globally at the top
quarters = ["Q1", "Q2", "Q3", "Q4"]

# --- STATE MANAGEMENT ---
if 'revenue_items' not in st.session_state:
    st.session_state.revenue_items = [{'name': 'Subscription Revenue', 'value': 50000.0, 'growth': 0.05}]

if 'cogs_items' not in st.session_state:
    st.session_state.cogs_items = [{'name': 'Hosting Costs', 'value': 0.20, 'type': '% of Rev'}]

if 'opex_items' not in st.session_state:
    st.session_state.opex_items = [{'name': 'Salaries', 'value': 15000.0, 'type': 'Fixed Amount', 'param2': 0.0}] 

if 'capex_items' not in st.session_state:
    st.session_state.capex_items = [{'name': 'Laptops', 'cost': 2000.0, 'deprec_rate': 0.20}]

if 'wc_assumptions' not in st.session_state:
    st.session_state.wc_assumptions = {
        'beginning_cash': 10000.0,
        'ar_percent': 0.10, 
        'ap_percent': 0.10,
        'deferred_rev_percent': 0.0
    }

if 'tax_assumptions' not in st.session_state:
    st.session_state.tax_assumptions = {
        'tax_rate': 0.25,
        'payment_timing': 'Immediate' 
    }

if 'financing_assumptions' not in st.session_state:
    st.session_state.financing_assumptions = {
        'equity_raised': 0.0,
        'debt_issued': 0.0,
        'debt_interest_rate': 0.05,
        'cash_interest_rate': 0.01
    }

# --- HELPER FUNCTIONS ---
def format_currency(val):
    return f"${val:,.0f}"

def format_percent(val):
    return f"{val*100:.1f}%"

# --- EXCEL GENERATION FUNCTION (Defined early for button) ---
def generate_excel():
    wb = Workbook()
    
    # 1. Assumptions Sheet
    ws_assump = wb.active
    ws_assump.title = "Assumptions"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    bold_font = Font(bold=True)
    currency_fmt = '#,##0'
    pct_fmt = '0.0%'
    
    ws_assump.append(["Category", "Driver", "Value", "Notes"])
    for cell in ws_assump[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    row_idx = 2
    refs = {} # Store cell references: refs['tax_rate'] = "Assumptions!C5"
    
    def add_assump(category, driver, value, fmt=None, key=None):
        nonlocal row_idx
        ws_assump.cell(row=row_idx, column=1, value=category)
        ws_assump.cell(row=row_idx, column=2, value=driver)
        c = ws_assump.cell(row=row_idx, column=3, value=value)
        if fmt: c.number_format = fmt
        ref = f"Assumptions!$C${row_idx}"
        if key: refs[key] = ref
        row_idx += 1
        return ref

    # Global Assumptions
    add_assump("Global", "Tax Rate", st.session_state.tax_assumptions['tax_rate'], pct_fmt, 'tax_rate')
    add_assump("Global", "Tax Payment (0=Imm, 1=NextYr)", 0 if st.session_state.tax_assumptions['payment_timing'] == "Immediate" else 1, None, 'tax_timing')
    add_assump("Working Capital", "Beginning Cash", st.session_state.wc_assumptions['beginning_cash'], currency_fmt, 'beg_cash')
    add_assump("Working Capital", "AR % of Rev", st.session_state.wc_assumptions['ar_percent'], pct_fmt, 'ar_pct')
    add_assump("Working Capital", "AP % of OpEx", st.session_state.wc_assumptions['ap_percent'], pct_fmt, 'ap_pct')
    add_assump("Working Capital", "Deferred Rev %", st.session_state.wc_assumptions.get('deferred_rev_percent', 0.0), pct_fmt, 'dr_pct')
    add_assump("Financing", "Equity Raised", st.session_state.financing_assumptions['equity_raised'], currency_fmt, 'equity')
    add_assump("Financing", "Debt Issued", st.session_state.financing_assumptions['debt_issued'], currency_fmt, 'debt')
    add_assump("Financing", "Debt Interest Rate", st.session_state.financing_assumptions['debt_interest_rate'], pct_fmt, 'debt_int')
    add_assump("Financing", "Cash Interest Rate", st.session_state.financing_assumptions['cash_interest_rate'], pct_fmt, 'cash_int')
    
    # Item Drivers
    refs['revenue'] = {}
    for item in st.session_state.revenue_items:
        refs['revenue'][item['name']] = {}
        refs['revenue'][item['name']]['start'] = add_assump("Revenue", f"{item['name']} - Start Value", item['value'], currency_fmt)
        refs['revenue'][item['name']]['growth'] = add_assump("Revenue", f"{item['name']} - Growth", item['growth'], pct_fmt)
        
    refs['cogs'] = {}
    for item in st.session_state.cogs_items:
        refs['cogs'][item['name']] = {}
        if item['type'] == "% of Rev":
            refs['cogs'][item['name']]['val'] = add_assump("COGS", f"{item['name']} - % of Rev", item['value'], pct_fmt)
        else:
            refs['cogs'][item['name']]['val'] = add_assump("COGS", f"{item['name']} - Fixed Amt", item['value'], currency_fmt)
            
    refs['opex'] = {}
    for item in st.session_state.opex_items:
        refs['opex'][item['name']] = {}
        if item['type'] == "Fixed Amount":
            refs['opex'][item['name']]['val'] = add_assump("OpEx", f"{item['name']} - Start Value", item['value'], currency_fmt)
            refs['opex'][item['name']]['growth'] = add_assump("OpEx", f"{item['name']} - Growth", item.get('param2', 0.0), pct_fmt)
        elif item['type'] == "% of Rev":
            refs['opex'][item['name']]['val'] = add_assump("OpEx", f"{item['name']} - % of Rev", item['value'], pct_fmt)
        elif item['type'] == "Personnel":
            refs['opex'][item['name']]['count'] = add_assump("OpEx", f"{item['name']} - Headcount", item['value'], None)
            refs['opex'][item['name']]['salary'] = add_assump("OpEx", f"{item['name']} - Avg Salary", item.get('param2', 0.0), currency_fmt)

    refs['capex'] = {}
    for item in st.session_state.capex_items:
        refs['capex'][item['name']] = {}
        refs['capex'][item['name']]['cost'] = add_assump("CapEx", f"{item['name']} - Cost", item['cost'], currency_fmt)
        refs['capex'][item['name']]['rate'] = add_assump("CapEx", f"{item['name']} - Deprec Rate", item.get('deprec_rate', 0.20), pct_fmt)

    ws_assump.column_dimensions['A'].width = 20
    ws_assump.column_dimensions['B'].width = 30
    ws_assump.column_dimensions['C'].width = 15

    # 2. Main Sheet
    ws = wb.create_sheet("3 Statement Model")
    
    headers = ["Item"] + quarters
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 2
    
    # --- P&L ---
    ws.cell(row=row_idx, column=1, value="PROFIT & LOSS").font = bold_font
    row_idx += 1
    
    # Revenue
    ws.cell(row=row_idx, column=1, value="Revenue").font = bold_font
    row_idx += 1
    rev_start_row = row_idx
    for item in st.session_state.revenue_items:
        ws.cell(row=row_idx, column=1, value=item['name'])
        start_ref = refs['revenue'][item['name']]['start']
        growth_ref = refs['revenue'][item['name']]['growth']
        for i, q in enumerate(quarters):
            col_letter = get_column_letter(i+2)
            if i == 0:
                ws.cell(row=row_idx, column=i+2, value=f"={start_ref}").number_format = currency_fmt
            else:
                prev_col = get_column_letter(i+1)
                ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx}*(1+{growth_ref})").number_format = currency_fmt
        row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Total Revenue").font = bold_font
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{rev_start_row}:{col_letter}{row_idx-1})").number_format = currency_fmt
    total_rev_row = row_idx
    row_idx += 2
    
    # COGS
    ws.cell(row=row_idx, column=1, value="Cost of Goods Sold").font = bold_font
    row_idx += 1
    cogs_start_row = row_idx
    for item in st.session_state.cogs_items:
        ws.cell(row=row_idx, column=1, value=item['name'])
        val_ref = refs['cogs'][item['name']]['val']
        for i, q in enumerate(quarters):
            col_letter = get_column_letter(i+2)
            if item['type'] == "% of Rev":
                ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_rev_row}*{val_ref}").number_format = currency_fmt
            else:
                ws.cell(row=row_idx, column=i+2, value=f"={val_ref}").number_format = currency_fmt
        row_idx += 1
        
    ws.cell(row=row_idx, column=1, value="Total COGS").font = bold_font
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{cogs_start_row}:{col_letter}{row_idx-1})").number_format = currency_fmt
    total_cogs_row = row_idx
    row_idx += 1
    
    # Gross Profit
    ws.cell(row=row_idx, column=1, value="Gross Profit").font = bold_font
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_rev_row}-{col_letter}{total_cogs_row}").number_format = currency_fmt
    gross_profit_row = row_idx
    row_idx += 2
    
    # OpEx
    ws.cell(row=row_idx, column=1, value="Operating Expenses").font = bold_font
    row_idx += 1
    opex_start_row = row_idx
    for item in st.session_state.opex_items:
        ws.cell(row=row_idx, column=1, value=item['name'])
        for i, q in enumerate(quarters):
            col_letter = get_column_letter(i+2)
            if item['type'] == "Fixed Amount":
                start_ref = refs['opex'][item['name']]['val']
                growth_ref = refs['opex'][item['name']]['growth']
                if i == 0:
                    ws.cell(row=row_idx, column=i+2, value=f"={start_ref}").number_format = currency_fmt
                else:
                    prev_col = get_column_letter(i+1)
                    ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx}*(1+{growth_ref})").number_format = currency_fmt
            elif item['type'] == "% of Rev":
                val_ref = refs['opex'][item['name']]['val']
                ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_rev_row}*{val_ref}").number_format = currency_fmt
            elif item['type'] == "Personnel":
                count_ref = refs['opex'][item['name']]['count']
                sal_ref = refs['opex'][item['name']]['salary']
                ws.cell(row=row_idx, column=i+2, value=f"=({count_ref}*{sal_ref})/4").number_format = currency_fmt
        row_idx += 1
        
    ws.cell(row=row_idx, column=1, value="Total Opex").font = bold_font
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{opex_start_row}:{col_letter}{row_idx-1})").number_format = currency_fmt
    total_opex_row = row_idx
    row_idx += 1
    
    # EBITDA
    ws.cell(row=row_idx, column=1, value="EBITDA").font = bold_font
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{gross_profit_row}-{col_letter}{total_opex_row}").number_format = currency_fmt
    ebitda_row = row_idx
    row_idx += 1
    
    # Depreciation
    ws.cell(row=row_idx, column=1, value="Depreciation")
    deprec_row = row_idx
    # Formula: Sum(Cost * Rate / 4) for all assets
    # Construct formula string: =(Cost1*Rate1/4) + (Cost2*Rate2/4) ...
    deprec_formula_parts = []
    for item in st.session_state.capex_items:
        cost_ref = refs['capex'][item['name']]['cost']
        rate_ref = refs['capex'][item['name']]['rate']
        deprec_formula_parts.append(f"({cost_ref}*{rate_ref}/4)")
    
    deprec_formula = "=" + "+".join(deprec_formula_parts) if deprec_formula_parts else "=0"
    
    for i, q in enumerate(quarters):
        ws.cell(row=row_idx, column=i+2, value=deprec_formula).number_format = currency_fmt
    row_idx += 1
    
    # EBIT
    ws.cell(row=row_idx, column=1, value="EBIT").font = bold_font
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ebitda_row}-{col_letter}{deprec_row}").number_format = currency_fmt
    ebit_row = row_idx
    row_idx += 1
    
    # Interest
    ws.cell(row=row_idx, column=1, value="Interest Expense")
    int_exp_row = row_idx
    for i, q in enumerate(quarters):
        ws.cell(row=row_idx, column=i+2, value=f"={refs['debt']}*{refs['debt_int']}/4").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Interest Income")
    int_inc_row = row_idx
    int_inc_cells = [] # Store cells to update later
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
            cell = ws.cell(row=row_idx, column=i+2, value=f"={refs['beg_cash']}*{refs['cash_int']}/4")
        else:
            prev_col = get_column_letter(i+1)
            # Use placeholder {CASH_ROW}
            cell = ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{{CASH_ROW}}*{refs['cash_int']}/4")
        cell.number_format = currency_fmt
        int_inc_cells.append(cell)
    row_idx += 1
    
    # EBT
    ws.cell(row=row_idx, column=1, value="EBT").font = bold_font
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ebit_row}-{col_letter}{int_exp_row}+{col_letter}{int_inc_row}").number_format = currency_fmt
    ebt_row = row_idx
    row_idx += 1
    
    # Taxes
    ws.cell(row=row_idx, column=1, value="Income Tax")
    tax_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=MAX(0, {col_letter}{ebt_row}*{refs['tax_rate']})").number_format = currency_fmt
    row_idx += 1
    
    # Net Income
    ws.cell(row=row_idx, column=1, value="Net Income").font = bold_font
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ebt_row}-{col_letter}{tax_row}").number_format = currency_fmt
    ni_row = row_idx
    row_idx += 3
    
    # --- BALANCE SHEET ---
    ws.cell(row=row_idx, column=1, value="BALANCE SHEET").font = bold_font
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Assets").font = bold_font
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Cash & Equivalents")
    cash_row = row_idx
    
    # Update Interest Income cells now that we know Cash Row
    for cell in int_inc_cells:
        if "{CASH_ROW}" in str(cell.value):
            cell.value = cell.value.replace("{CASH_ROW}", str(cash_row))
            
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Accounts Receivable")
    ar_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_rev_row}*{refs['ar_pct']}").number_format = currency_fmt
    row_idx += 1
    
    # Fixed Assets
    ws.cell(row=row_idx, column=1, value="Fixed Assets (Gross)")
    fa_row = row_idx
    fa_cells = [] # Store cells to update later
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        # Use placeholder {CAPEX_ROW}
        if i == 0:
            cell = ws.cell(row=row_idx, column=i+2, value=f"=-{col_letter}{{CAPEX_ROW}}")
        else:
            prev_col = get_column_letter(i+1)
            cell = ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx}-{col_letter}{{CAPEX_ROW}}")
        cell.number_format = currency_fmt
        fa_cells.append(cell)
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Accumulated Depreciation")
    ad_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"=-{col_letter}{deprec_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx}-{col_letter}{deprec_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Total Assets").font = bold_font
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{cash_row}:{col_letter}{ad_row})").number_format = currency_fmt
    row_idx += 2
    
    ws.cell(row=row_idx, column=1, value="Liabilities & Equity").font = bold_font
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Accounts Payable")
    ap_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_opex_row}*{refs['ap_pct']}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Deferred Revenue")
    dr_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_rev_row}*{refs['dr_pct']}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Tax Payable")
    tp_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"=IF({refs['tax_timing']}=0, 0, {col_letter}{tax_row})").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"=IF({refs['tax_timing']}=0, 0, {col_letter}{tax_row})").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Long Term Debt")
    debt_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={refs['debt']}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Common Stock")
    cs_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={refs['beg_cash']}+{refs['equity']}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Retained Earnings")
    re_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ni_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{re_row}+{col_letter}{ni_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Total Liab & Equity").font = bold_font
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{ap_row}:{col_letter}{re_row})").number_format = currency_fmt
    row_idx += 3
    
    # --- CASH FLOW ---
    ws.cell(row=row_idx, column=1, value="CASH FLOW STATEMENT").font = bold_font
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Cash from Operations").font = bold_font
    row_idx += 1
    cfo_start_row = row_idx
    
    ws.cell(row=row_idx, column=1, value="Net Income")
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ni_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Depreciation")
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{deprec_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Change in AR")
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"=-{col_letter}{ar_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{ar_row}-{col_letter}{ar_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Change in AP")
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ap_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ap_row}-{prev_col}{ap_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Change in Deferred Rev")
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{dr_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{dr_row}-{prev_col}{dr_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Change in Tax Payable")
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{tp_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{tp_row}-{prev_col}{tp_row}").number_format = currency_fmt
    row_idx += 1
    cfo_end_row = row_idx - 1
    
    ws.cell(row=row_idx, column=1, value="Cash from Investing").font = bold_font
    row_idx += 1
    cfi_start_row = row_idx
    
    ws.cell(row=row_idx, column=1, value="CapEx")
    
    # Update Fixed Assets formulas now that we know the row
    for cell in fa_cells:
        cell.value = cell.value.replace("{CAPEX_ROW}", str(row_idx))
    
    # Formula: Sum of Cost Assumptions.
    capex_formula_parts = []
    for item in st.session_state.capex_items:
        capex_formula_parts.append(refs['capex'][item['name']]['cost'])
    # Remove leading '=' because we prepend '=-' in the cell value
    capex_formula = "(" + "+".join(capex_formula_parts) + ")" if capex_formula_parts else "0"
    
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"=-{capex_formula}").number_format = currency_fmt
        else:
            ws.cell(row=row_idx, column=i+2, value=0).number_format = currency_fmt
    row_idx += 1
    cfi_end_row = row_idx - 1
    
    ws.cell(row=row_idx, column=1, value="Cash from Financing").font = bold_font
    row_idx += 1
    cff_start_row = row_idx
    
    ws.cell(row=row_idx, column=1, value="Issuance of Common Stock")
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
             ws.cell(row=row_idx, column=i+2, value=f"={refs['beg_cash']}+{refs['equity']}").number_format = currency_fmt
        else:
             ws.cell(row=row_idx, column=i+2, value=0).number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Issuance of Debt")
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
             ws.cell(row=row_idx, column=i+2, value=f"={refs['debt']}").number_format = currency_fmt
        else:
             ws.cell(row=row_idx, column=i+2, value=0).number_format = currency_fmt
    row_idx += 1
    cff_end_row = row_idx - 1
    
    ws.cell(row=row_idx, column=1, value="Net Cash Flow").font = bold_font
    ncf_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{cfo_start_row}:{col_letter}{cfo_end_row})+SUM({col_letter}{cfi_start_row}:{col_letter}{cfi_end_row})+SUM({col_letter}{cff_start_row}:{col_letter}{cff_end_row})").number_format = currency_fmt
    row_idx += 2
    
    ws.cell(row=row_idx, column=1, value="Ending Cash Balance").font = bold_font
    ec_row = row_idx
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        if i == 0:
             ws.cell(row=row_idx, column=i+2, value=f"=0+{col_letter}{ncf_row}").number_format = currency_fmt
        else:
             prev_col = get_column_letter(i+1)
             ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx}+{col_letter}{ncf_row}").number_format = currency_fmt
             
    # Link BS Cash
    for i, q in enumerate(quarters):
        col_letter = get_column_letter(i+2)
        ws.cell(row=cash_row, column=i+2, value=f"={col_letter}{ec_row}").number_format = currency_fmt
    
    ws.column_dimensions['A'].width = 30
    for col in ['B','C','D','E']:
        ws.column_dimensions[col].width = 15
        
    return wb

# --- DOWNLOAD BUTTON (TOP OF PAGE) ---
try:
    wb = generate_excel()
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    col_dl1, col_dl2 = st.columns([3, 1])
    with col_dl2:
        st.download_button(
            label="📥 Download Dynamic Model (.xlsx)",
            data=buffer,
            file_name="Dynamic_Financial_Model.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
except Exception as e:
    st.error(f"Error generating Excel file: {e}")

# --- MAIN UI INPUTS ---
st.markdown("### Model Assumptions")

col_main1, col_main2 = st.columns(2)

with col_main1:
    # 1. Revenue Streams
    with st.expander("1. Revenue Streams", expanded=True):
        for i, item in enumerate(st.session_state.revenue_items):
            st.markdown(f"**{item['name']}**")
            c1, c2 = st.columns(2)
            item['value'] = c1.number_input(f"Q1 Value ($) ##{i}", value=item['value'], step=1000.0, key=f"rev_val_{i}")
            item['growth'] = c2.number_input(f"Growth Rate (%) ##{i}", value=item['growth']*100, step=1.0, key=f"rev_gr_{i}") / 100
            if st.button(f"Remove {item['name']}", key=f"del_rev_{i}"):
                st.session_state.revenue_items.pop(i)
                st.rerun()
        
        st.markdown("---")
        new_rev_name = st.text_input("New Revenue Name", key="new_rev_name")
        if st.button("Add Revenue Stream"):
            if new_rev_name:
                st.session_state.revenue_items.append({'name': new_rev_name, 'value': 0.0, 'growth': 0.0})
                st.rerun()

    # 3. OpEx
    with st.expander("3. OpEx (People vs Fixed)", expanded=False):
        for i, item in enumerate(st.session_state.opex_items):
            st.markdown(f"**{item['name']}**")
            type_opts = ["Fixed Amount", "% of Rev", "Personnel"]
            curr_type_idx = type_opts.index(item.get('type', 'Fixed Amount'))
            item['type'] = st.selectbox(f"Type ##{i}", type_opts, index=curr_type_idx, key=f"opex_type_{i}")
            
            c1, c2 = st.columns(2)
            if item['type'] == "Fixed Amount":
                item['value'] = c1.number_input(f"Fixed Amount ($) ##{i}", value=float(item['value']), step=1000.0, key=f"opex_val_{i}")
                item['param2'] = c2.number_input(f"Growth (%) ##{i}", value=float(item.get('param2', 0.0))*100, step=1.0, key=f"opex_p2_{i}") / 100
            elif item['type'] == "% of Rev":
                item['value'] = c1.number_input(f"% of Revenue ##{i}", value=float(item['value'])*100, step=1.0, key=f"opex_val_{i}") / 100
                item['param2'] = 0.0
            elif item['type'] == "Personnel":
                item['value'] = c1.number_input(f"Headcount (Start) ##{i}", value=float(item['value']), step=1.0, key=f"opex_val_{i}")
                item['param2'] = c2.number_input(f"Avg Salary ($/yr) ##{i}", value=float(item.get('param2', 50000.0)), step=1000.0, key=f"opex_p2_{i}")

            if st.button(f"Remove {item['name']}", key=f"del_opex_{i}"):
                st.session_state.opex_items.pop(i)
                st.rerun()

        st.markdown("---")
        new_opex_name = st.text_input("New OpEx Name", key="new_opex_name")
        if st.button("Add OpEx Item"):
            if new_opex_name:
                st.session_state.opex_items.append({'name': new_opex_name, 'value': 5000.0, 'type': 'Fixed Amount', 'param2': 0.0})
                st.rerun()

    # 5. Assets & Taxation
    with st.expander("5. Assets & Taxation", expanded=False):
        st.markdown("**Assets**")
        for i, item in enumerate(st.session_state.capex_items):
            st.markdown(f"**{item['name']}**")
            c1, c2 = st.columns(2)
            item['cost'] = c1.number_input(f"Cost ($) ##{i}", value=item['cost'], step=500.0, key=f"capex_cost_{i}")
            item['deprec_rate'] = c2.number_input(f"Deprec Rate (%) ##{i}", value=item.get('deprec_rate', 0.20)*100, step=1.0, key=f"capex_rate_{i}") / 100
            if st.button(f"Remove {item['name']}", key=f"del_capex_{i}"):
                st.session_state.capex_items.pop(i)
                st.rerun()
        
        new_capex_name = st.text_input("New Asset Name", key="new_capex_name")
        if st.button("Add Asset"):
            if new_capex_name:
                st.session_state.capex_items.append({'name': new_capex_name, 'cost': 1000.0, 'deprec_rate': 0.20})
                st.rerun()
        
        st.markdown("---")
        st.markdown("**Taxation**")
        st.session_state.tax_assumptions['tax_rate'] = st.slider("Tax Rate (%)", 0, 50, int(st.session_state.tax_assumptions['tax_rate']*100)) / 100
        st.session_state.tax_assumptions['payment_timing'] = st.radio("Tax Payment Timing", ["Immediate", "Next Year"], index=0 if st.session_state.tax_assumptions['payment_timing'] == "Immediate" else 1)


with col_main2:
    # 2. COGS
    with st.expander("2. COGS", expanded=False):
        for i, item in enumerate(st.session_state.cogs_items):
            st.markdown(f"**{item['name']}**")
            c1, c2 = st.columns(2)
            item['type'] = c1.selectbox(f"Type ##{i}", ["% of Rev", "Fixed Amount"], index=0 if item['type'] == "% of Rev" else 1, key=f"cogs_type_{i}")
            
            if item['type'] == "% of Rev":
                val = c2.number_input(f"% of Rev ##{i}", value=item['value']*100, step=1.0, key=f"cogs_val_{i}") / 100
            else:
                val = c2.number_input(f"Fixed ($) ##{i}", value=item['value'], step=500.0, key=f"cogs_val_{i}")
            item['value'] = val
            
            if st.button(f"Remove {item['name']}", key=f"del_cogs_{i}"):
                st.session_state.cogs_items.pop(i)
                st.rerun()

        st.markdown("---")
        new_cogs_name = st.text_input("New COGS Name", key="new_cogs_name")
        if st.button("Add COGS Item"):
            if new_cogs_name:
                st.session_state.cogs_items.append({'name': new_cogs_name, 'value': 0.10, 'type': '% of Rev'})
                st.rerun()

    # 4. Working Capital
    with st.expander("4. Working Capital", expanded=False):
        st.session_state.wc_assumptions['beginning_cash'] = st.number_input("Beginning Cash Balance ($)", value=st.session_state.wc_assumptions['beginning_cash'], step=1000.0)
        st.session_state.wc_assumptions['ar_percent'] = st.slider("AR as % of Rev (DSO Proxy)", 0, 50, int(st.session_state.wc_assumptions['ar_percent']*100)) / 100
        st.session_state.wc_assumptions['ap_percent'] = st.slider("AP as % of Opex (DPO Proxy)", 0, 50, int(st.session_state.wc_assumptions['ap_percent']*100)) / 100
        st.session_state.wc_assumptions['deferred_rev_percent'] = st.slider("Deferred Revenue (% of Sales)", 0, 50, int(st.session_state.wc_assumptions.get('deferred_rev_percent', 0.0)*100)) / 100

    # 6. Financing
    with st.expander("6. Financing", expanded=False):
        st.session_state.financing_assumptions['equity_raised'] = st.number_input("New Equity Raised ($)", value=st.session_state.financing_assumptions['equity_raised'], step=1000.0)
        st.session_state.financing_assumptions['debt_issued'] = st.number_input("New Debt Issued ($)", value=st.session_state.financing_assumptions['debt_issued'], step=1000.0)
        st.session_state.financing_assumptions['debt_interest_rate'] = st.number_input("Interest Rate on Debt (%)", value=st.session_state.financing_assumptions['debt_interest_rate']*100, step=0.1) / 100
        st.session_state.financing_assumptions['cash_interest_rate'] = st.number_input("Interest on Cash (%)", value=st.session_state.financing_assumptions['cash_interest_rate']*100, step=0.1) / 100

st.markdown("---")