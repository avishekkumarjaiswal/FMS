import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Set up the Streamlit page (must be the first command)
st.set_page_config(layout="wide")  # Use the full width of the screen

# Hide Streamlit menu, footer, and prevent code inspection
st.markdown("""
    <style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden; display: none;}
    .stDeployButton {display: none !important;}  /* Hide GitHub button */
    </style>

    <script>
    document.addEventListener('contextmenu', event => event.preventDefault());
    document.onkeydown = function(e) {
        if (e.ctrlKey && (e.keyCode === 85 || e.keyCode === 83)) {
            return false;  // Disable "Ctrl + U" (View Source) & "Ctrl + S" (Save As)
        }
        if (e.keyCode == 123) {
            return false;  // Disable "F12" (DevTools)
        }
    };
    </script>
    """, unsafe_allow_html=True)

# --- GLOBAL VARIABLES ---
SCENARIOS = ["Base", "Optimistic", "Pessimistic"]
periods = [f"Month {i+1}" for i in range(36)]

# --- SESSION STATE INITIALIZATION ---
if 'scenario_to_edit' not in st.session_state:
    st.session_state.scenario_to_edit = "Base"
if 'scenario_to_run' not in st.session_state:
    st.session_state.scenario_to_run = "Base"

def init_scenario_val(val):
    return {s: val for s in SCENARIOS}

if 'revenue_items' not in st.session_state:
    st.session_state.revenue_items = [
        {'name': 'Product Sales', 'value': init_scenario_val(100000.0), 
         'growth_y1': init_scenario_val(0.10), 'growth_y2': init_scenario_val(0.07), 'growth_y3': init_scenario_val(0.04)},
        {'name': 'Service Revenue', 'value': init_scenario_val(50000.0), 
         'growth_y1': init_scenario_val(0.05), 'growth_y2': init_scenario_val(0.03), 'growth_y3': init_scenario_val(0.02)}
    ]
    
if 'cogs_items' not in st.session_state:
    st.session_state.cogs_items = [
        {'name': 'Hosting Costs', 'value': init_scenario_val(0.20), 'type': '% of Rev'},
        {'name': 'Support Staff', 'value': init_scenario_val(20000.0), 'type': 'Fixed Amount'}
    ]

if 'opex_items' not in st.session_state:
    st.session_state.opex_items = [
        {'name': 'Marketing', 'value': init_scenario_val(10000.0), 'type': 'Fixed Amount', 'param2': init_scenario_val(0.05)},
        {'name': 'Sales Team', 'value': init_scenario_val(2.0), 'type': 'Personnel', 'param2': init_scenario_val(60000.0), 'revenue_threshold': init_scenario_val(50000.0)}
    ]

if 'capex_items' not in st.session_state:
    st.session_state.capex_items = [
        {'name': 'Servers', 'cost': init_scenario_val(50000.0), 'deprec_rate': init_scenario_val(0.20)},
        {'name': 'Laptops', 'cost': init_scenario_val(10000.0), 'deprec_rate': init_scenario_val(0.33)}
    ]

if 'tax_assumptions' not in st.session_state:
    st.session_state.tax_assumptions = {
        'tax_rate': init_scenario_val(0.25),
        'payment_timing': 'Immediate',
        'nol_balance': 0.0
    }

if 'wc_assumptions' not in st.session_state:
    st.session_state.wc_assumptions = {
        'beginning_cash': 50000.0,
        'ar_percent': init_scenario_val(0.10),
        'ap_percent': init_scenario_val(0.10),
        'deferred_rev_percent': init_scenario_val(0.0),
        'days_inventory': init_scenario_val(30.0),
        'days_payable': init_scenario_val(30.0)
    }

if 'financing_assumptions' not in st.session_state:
    st.session_state.financing_assumptions = {
        'equity_raised': init_scenario_val(0.0),
        'debt_issued': init_scenario_val(0.0),
        'debt_interest_rate': init_scenario_val(0.05),
        'cash_interest_rate': init_scenario_val(0.02),
        'overdraft_interest_rate': init_scenario_val(0.10),
        'debt_repayment_term': init_scenario_val(5)
    }

if 'capex_assumptions' not in st.session_state:
    st.session_state.capex_assumptions = {
        'maintenance_pct': init_scenario_val(0.02)  # 2% of revenue per month
    }

if 'kpi_assumptions' not in st.session_state:
    st.session_state.kpi_assumptions = {
        'starting_customers': init_scenario_val(100.0),  # Initial customer count
        'new_customers_monthly': init_scenario_val(10.0),  # New customers per month
        'churn_rate_monthly': init_scenario_val(0.02),  # 2% monthly churn
        'sm_opex_items': ['Marketing', 'Sales Team']  # Default S&M items
    }

# --- EXCEL GENERATION FUNCTION ---
def generate_excel():
    wb = Workbook()
    # Enable Iterative Calculation for Circular References
    wb.calculation.iterate = True
    wb.calculation.iterateCount = 100
    wb.calculation.iterateDelta = 0.001
    
    scen = st.session_state.scenario_to_run
    
    # 1. Assumptions Sheet
    ws_assump = wb.active
    ws_assump.title = "Assumptions"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    bold_font = Font(bold=True)
    currency_fmt = '#,##0.00'
    pct_fmt = '0.00%'
    
    ws_assump.append(["Category", "Driver", "Value", "Notes"])
    for cell in ws_assump[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    row_idx = 2
    refs = {} 
    
    def add_assump(category, driver, value, fmt=None, key=None):
        nonlocal row_idx
        ws_assump.cell(row=row_idx, column=1, value=category)
        ws_assump.cell(row=row_idx, column=2, value=driver)
        c = ws_assump.cell(row=row_idx, column=3, value=value)
        if fmt: c.number_format = fmt
        ref = f"Assumptions!$C${row_idx}"
        if key: refs[key] = ref
        row_idx += 1
        return ref

    # Global Assumptions
    add_assump("Global", "Tax Rate", st.session_state.tax_assumptions['tax_rate'][scen], pct_fmt, 'tax_rate')
    add_assump("Global", "Tax Payment (0=Imm, 1=NextYr)", 0 if st.session_state.tax_assumptions['payment_timing'] == "Immediate" else 1, None, 'tax_timing')
    refs['beg_nol'] = add_assump("Global", "NOL Beginning Balance", st.session_state.tax_assumptions['nol_balance'], currency_fmt)
    
    # Working Capital
    add_assump("Working Capital", "Beginning Cash", st.session_state.wc_assumptions['beginning_cash'], currency_fmt, 'beg_cash')
    add_assump("Working Capital", "AR % of Rev", st.session_state.wc_assumptions['ar_percent'][scen], pct_fmt, 'ar_pct')
    add_assump("Working Capital", "AP % of OpEx", st.session_state.wc_assumptions['ap_percent'][scen], pct_fmt, 'ap_pct')
    add_assump("Working Capital", "Deferred Rev %", st.session_state.wc_assumptions['deferred_rev_percent'][scen], pct_fmt, 'dr_pct')
    add_assump("Working Capital", "Days Inventory Outstanding (DIO)", st.session_state.wc_assumptions['days_inventory'][scen], None, 'dio')
    add_assump("Working Capital", "Days Payable Outstanding (DPO)", st.session_state.wc_assumptions['days_payable'][scen], None, 'dpo')
    
    # Financing
    add_assump("Financing", "Equity Raised", st.session_state.financing_assumptions['equity_raised'][scen], currency_fmt, 'equity')
    add_assump("Financing", "Debt Issued", st.session_state.financing_assumptions['debt_issued'][scen], currency_fmt, 'debt')
    add_assump("Financing", "Debt Interest Rate", st.session_state.financing_assumptions['debt_interest_rate'][scen], pct_fmt, 'debt_int')
    add_assump("Financing", "Cash Interest Rate", st.session_state.financing_assumptions['cash_interest_rate'][scen], pct_fmt, 'cash_int')
    add_assump("Financing", "Overdraft Interest Rate", st.session_state.financing_assumptions['overdraft_interest_rate'][scen], pct_fmt, 'od_int')
    add_assump("Financing", "Debt Repayment Term (Years)", st.session_state.financing_assumptions['debt_repayment_term'][scen], None, 'debt_term')

    refs['revenue'] = {}
    for item in st.session_state.revenue_items:
        refs['revenue'][item['name']] = {}
        refs['revenue'][item['name']]['start'] = add_assump("Revenue", f"{item['name']} - Start Value", item['value'][scen], currency_fmt)
        refs['revenue'][item['name']]['growth_y1'] = add_assump("Revenue", f"{item['name']} - Y1 Growth", item.get('growth_y1', {}).get(scen, 0.10), pct_fmt)
        refs['revenue'][item['name']]['growth_y2'] = add_assump("Revenue", f"{item['name']} - Y2 Growth", item.get('growth_y2', {}).get(scen, 0.07), pct_fmt)
        refs['revenue'][item['name']]['growth_y3'] = add_assump("Revenue", f"{item['name']} - Y3 Growth", item.get('growth_y3', {}).get(scen, 0.04), pct_fmt)
        
    refs['cogs'] = {}
    for item in st.session_state.cogs_items:
        refs['cogs'][item['name']] = {}
        if item['type'] == "% of Rev":
            refs['cogs'][item['name']]['val'] = add_assump("COGS", f"{item['name']} - % of Rev", item['value'][scen], pct_fmt)
        else:
            refs['cogs'][item['name']]['val'] = add_assump("COGS", f"{item['name']} - Fixed Amt", item['value'][scen], currency_fmt)
            
    refs['opex'] = {}
    for item in st.session_state.opex_items:
        refs['opex'][item['name']] = {}
        if item['type'] == "Fixed Amount":
            refs['opex'][item['name']]['val'] = add_assump("OpEx", f"{item['name']} - Start Value", item['value'][scen], currency_fmt)
            refs['opex'][item['name']]['growth'] = add_assump("OpEx", f"{item['name']} - Growth", item.get('param2', {}).get(scen, 0.0), pct_fmt)
        elif item['type'] == "% of Rev":
            refs['opex'][item['name']]['val'] = add_assump("OpEx", f"{item['name']} - % of Rev", item['value'][scen], pct_fmt)
        elif item['type'] == "Personnel":
            refs['opex'][item['name']]['count'] = add_assump("OpEx", f"{item['name']} - Headcount", item['value'][scen], None)
            refs['opex'][item['name']]['salary'] = add_assump("OpEx", f"{item['name']} - Avg Salary", item.get('param2', {}).get(scen, 0.0), currency_fmt)
            if item.get('revenue_threshold'):
                refs['opex'][item['name']]['threshold'] = add_assump("OpEx", f"{item['name']} - Revenue Threshold ($)", item.get('revenue_threshold', {}).get(scen, 50000.0), currency_fmt)

    refs['capex'] = {}
    for item in st.session_state.capex_items:
        refs['capex'][item['name']] = {}
        refs['capex'][item['name']]['cost'] = add_assump("CapEx", f"{item['name']} - Cost", item['cost'][scen], currency_fmt)
        refs['capex'][item['name']]['rate'] = add_assump("CapEx", f"{item['name']} - Deprec Rate", item.get('deprec_rate', {}).get(scen, 0.20), pct_fmt)
    
    # Maintenance CapEx
    add_assump("CapEx", "Maintenance CapEx (% of Revenue)", st.session_state.capex_assumptions.get('maintenance_pct', {}).get(scen, 0.02), pct_fmt, 'maint_capex')

    ws_assump.column_dimensions['A'].width = 20
    ws_assump.column_dimensions['B'].width = 30
    ws_assump.column_dimensions['C'].width = 15

    # 2. Main Sheet
    ws = wb.create_sheet("36 Month Model")
    
    headers = ["Item"] + periods
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    row_idx = 2
    
    # --- P&L ---
    ws.cell(row=row_idx, column=1, value="PROFIT & LOSS").font = bold_font
    row_idx += 1
    
    # Revenue
    ws.cell(row=row_idx, column=1, value="Revenue").font = bold_font
    row_idx += 1
    rev_start_row = row_idx
    for item in st.session_state.revenue_items:
        ws.cell(row=row_idx, column=1, value=item['name'])
        start_ref = refs['revenue'][item['name']]['start']
        growth_y1 = refs['revenue'][item['name']]['growth_y1']
        growth_y2 = refs['revenue'][item['name']]['growth_y2']
        growth_y3 = refs['revenue'][item['name']]['growth_y3']
        for i, p in enumerate(periods):
            col_letter = get_column_letter(i+2)
            if i == 0:
                ws.cell(row=row_idx, column=i+2, value=f"={start_ref}/12").number_format = currency_fmt
            else:
                prev_col = get_column_letter(i+1)
                month_num = i + 1
                formula = f"={prev_col}{row_idx}*((1+IF({month_num}<=12,{growth_y1},IF({month_num}<=24,{growth_y2},{growth_y3})))^(1/12))"
                ws.cell(row=row_idx, column=i+2, value=formula).number_format = currency_fmt
        row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Total Revenue").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{rev_start_row}:{col_letter}{row_idx-1})").number_format = currency_fmt
    total_rev_row = row_idx
    row_idx += 1
    
    # COGS
    ws.cell(row=row_idx, column=1, value="Cost of Goods Sold").font = bold_font
    row_idx += 1
    cogs_start_row = row_idx
    for item in st.session_state.cogs_items:
        ws.cell(row=row_idx, column=1, value=item['name'])
        val_ref = refs['cogs'][item['name']]['val']
        for i, p in enumerate(periods):
            col_letter = get_column_letter(i+2)
            if item['type'] == "% of Rev":
                ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_rev_row}*{val_ref}").number_format = currency_fmt
            else:
                ws.cell(row=row_idx, column=i+2, value=f"={val_ref}/12").number_format = currency_fmt
        row_idx += 1
        
    ws.cell(row=row_idx, column=1, value="Total COGS").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{cogs_start_row}:{col_letter}{row_idx-1})").number_format = currency_fmt
    total_cogs_row = row_idx
    row_idx += 1
    
    # Gross Profit
    ws.cell(row=row_idx, column=1, value="Gross Profit").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_rev_row}-{col_letter}{total_cogs_row}").number_format = currency_fmt
    gross_profit_row = row_idx
    row_idx += 2
    
    # OpEx
    ws.cell(row=row_idx, column=1, value="Operating Expenses").font = bold_font
    row_idx += 1
    opex_start_row = row_idx
    for item in st.session_state.opex_items:
        ws.cell(row=row_idx, column=1, value=item['name'])
        for i, p in enumerate(periods):
            col_letter = get_column_letter(i+2)
            if item['type'] == "Fixed Amount":
                start_ref = refs['opex'][item['name']]['val']
                growth_ref = refs['opex'][item['name']]['growth']
                if i == 0:
                    ws.cell(row=row_idx, column=i+2, value=f"={start_ref}/12").number_format = currency_fmt
                else:
                    prev_col = get_column_letter(i+1)
                    ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx}*((1+{growth_ref})^(1/12))").number_format = currency_fmt
            elif item['type'] == "% of Rev":
                val_ref = refs['opex'][item['name']]['val']
                ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_rev_row}*{val_ref}").number_format = currency_fmt
            elif item['type'] == "Personnel":
                count_ref = refs['opex'][item['name']]['count']
                sal_ref = refs['opex'][item['name']]['salary']
                if item.get('revenue_threshold'):
                    thresh_ref = refs['opex'][item['name']].get('threshold', count_ref)
                    formula = f"=(({count_ref}+FLOOR(MAX(0,{col_letter}{total_rev_row}-B{total_rev_row})/{thresh_ref},1))*{sal_ref})/12"
                    ws.cell(row=row_idx, column=i+2, value=formula).number_format = currency_fmt
                else:
                    ws.cell(row=row_idx, column=i+2, value=f"=({count_ref}*{sal_ref})/12").number_format = currency_fmt
        row_idx += 1
        
    ws.cell(row=row_idx, column=1, value="Total Opex").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{opex_start_row}:{col_letter}{row_idx-1})").number_format = currency_fmt
    total_opex_row = row_idx
    row_idx += 1
    
    # EBITDA
    ws.cell(row=row_idx, column=1, value="EBITDA").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{gross_profit_row}-{col_letter}{total_opex_row}").number_format = currency_fmt
    ebitda_row = row_idx
    row_idx += 1
    
    # Depreciation
    ws.cell(row=row_idx, column=1, value="Depreciation")
    deprec_row = row_idx
    # Formula: Sum(Cost * Rate) / 12 for monthly
    deprec_formula_parts = []
    for item in st.session_state.capex_items:
        cost_ref = refs['capex'][item['name']]['cost']
        rate_ref = refs['capex'][item['name']]['rate']
        deprec_formula_parts.append(f"({cost_ref}*{rate_ref})")
    
    deprec_formula = "=(" + "+".join(deprec_formula_parts) + ")/12" if deprec_formula_parts else "=0"
    
    for i, p in enumerate(periods):
        ws.cell(row=row_idx, column=i+2, value=deprec_formula).number_format = currency_fmt
    row_idx += 1
    
    # EBIT
    ws.cell(row=row_idx, column=1, value="EBIT").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ebitda_row}-{col_letter}{deprec_row}").number_format = currency_fmt
    ebit_row = row_idx
    row_idx += 1
    
    # Interest Expense (Debt * Rate / 12)
    ws.cell(row=row_idx, column=1, value="Interest Expense (Debt)")
    int_exp_row = row_idx
    int_exp_cells = []
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        # Placeholder {DEBT_ROW}
        cell = ws.cell(row=row_idx, column=i+2, value=f"={{DEBT_ROW}}*{refs['debt_int']}/12")
        cell.number_format = currency_fmt
        int_exp_cells.append(cell)
    row_idx += 1

    # Overdraft Interest (Cash < 0)
    ws.cell(row=row_idx, column=1, value="Interest Expense (Overdraft)")
    od_int_row = row_idx
    od_int_cells = []
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        # Use Beginning Cash to avoid circular reference
        # IF(Beg_Cash < 0, ABS(Beg_Cash) * Rate/12, 0)
        if i == 0:
            cell = ws.cell(row=row_idx, column=i+2, value=f"=IF({refs['beg_cash']}<0, ABS({refs['beg_cash']})*{refs['od_int']}/12, 0)")
        else:
            prev_col = get_column_letter(i+1)
            cell = ws.cell(row=row_idx, column=i+2, value=f"=IF({prev_col}{{CASH_ROW}}<0, ABS({prev_col}{{CASH_ROW}})*{refs['od_int']}/12, 0)")
        cell.number_format = currency_fmt
        od_int_cells.append(cell)
    row_idx += 1
    
    # Interest Income (Cash > 0)
    ws.cell(row=row_idx, column=1, value="Interest Income")
    int_inc_row = row_idx
    int_inc_cells = []
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        # Use Beginning Cash to avoid circular reference
        # IF(Beg_Cash > 0, Beg_Cash * Rate/12, 0)
        if i == 0:
            cell = ws.cell(row=row_idx, column=i+2, value=f"=IF({refs['beg_cash']}>0, {refs['beg_cash']}*{refs['cash_int']}/12, 0)")
        else:
            prev_col = get_column_letter(i+1)
            cell = ws.cell(row=row_idx, column=i+2, value=f"=IF({prev_col}{{CASH_ROW}}>0, {prev_col}{{CASH_ROW}}*{refs['cash_int']}/12, 0)")
        cell.number_format = currency_fmt
        int_inc_cells.append(cell)
    row_idx += 1
    
    # EBT
    ws.cell(row=row_idx, column=1, value="EBT").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ebit_row}-{col_letter}{int_exp_row}-{col_letter}{od_int_row}+{col_letter}{int_inc_row}").number_format = currency_fmt
    ebt_row = row_idx
    row_idx += 1
    
    # Taxes (NOL Logic)
    ws.cell(row=row_idx, column=1, value="NOL Beginning Balance")
    nol_beg_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"={refs['beg_nol']}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx+2}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Taxable Income")
    taxable_inc_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=MAX(0, {col_letter}{ebt_row}-{col_letter}{nol_beg_row})").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="NOL Ending Balance")
    nol_end_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=IF({col_letter}{ebt_row}<0, {col_letter}{nol_beg_row}-{col_letter}{ebt_row}, MAX(0, {col_letter}{nol_beg_row}-{col_letter}{ebt_row}))").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Income Tax")
    tax_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{taxable_inc_row}*{refs['tax_rate']}").number_format = currency_fmt
    row_idx += 1
    
    # Net Income
    ws.cell(row=row_idx, column=1, value="Net Income").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ebt_row}-{col_letter}{tax_row}").number_format = currency_fmt
    ni_row = row_idx
    row_idx += 3
    
    # --- BALANCE SHEET ---
    ws.cell(row=row_idx, column=1, value="BALANCE SHEET").font = bold_font
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Assets").font = bold_font
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Cash & Equivalents")
    cash_row = row_idx
    # Update Interest Income cells
    for cell in int_inc_cells:
        if "{CASH_ROW}" in str(cell.value):
            cell.value = cell.value.replace("{CASH_ROW}", str(cash_row))
    # Update Overdraft Interest cells
    for cell in od_int_cells:
        if "{CASH_ROW}" in str(cell.value):
            cell.value = cell.value.replace("{CASH_ROW}", str(cash_row))
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Accounts Receivable")
    ar_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_rev_row}*{refs['ar_pct']}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Inventory")
    inv_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=({col_letter}{total_cogs_row}/30)*{refs['dio']}").number_format = currency_fmt # Approx 30 days/month
    row_idx += 1
    
    # Fixed Assets
    ws.cell(row=row_idx, column=1, value="Fixed Assets (Gross)")
    fa_row = row_idx
    fa_cells = []
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            cell = ws.cell(row=row_idx, column=i+2, value=f"=-{col_letter}{{CAPEX_ROW}}")
        else:
            prev_col = get_column_letter(i+1)
            cell = ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx}-{col_letter}{{CAPEX_ROW}}")
        cell.number_format = currency_fmt
        fa_cells.append(cell)
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Accumulated Depreciation")
    ad_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"=-{col_letter}{deprec_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx}-{col_letter}{deprec_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Total Assets").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{cash_row}:{col_letter}{ad_row})").number_format = currency_fmt
    row_idx += 2
    
    ws.cell(row=row_idx, column=1, value="Liabilities & Equity").font = bold_font
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Accounts Payable")
    ap_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=({col_letter}{total_cogs_row}/30)*{refs['dpo']}").number_format = currency_fmt # Approx 30 days/month
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Deferred Revenue")
    dr_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{total_rev_row}*{refs['dr_pct']}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Tax Payable")
    tp_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        # Simplified: If Immediate, 0. If Next Year, accumulate.
        # For monthly, "Next Year" is tricky. Let's assume "Deferred" means pay in Month 4 of next year or similar?
        # For simplicity in this model: Immediate = 0, Deferred = Accumulate and never pay (needs logic) or pay in Month 13?
        # Let's keep it simple: Deferred = Accumulate.
        ws.cell(row=row_idx, column=i+2, value=f"=IF({refs['tax_timing']}=0, 0, {col_letter}{tax_row})").number_format = currency_fmt
        # Note: This simple formula for deferred just sets it to current tax. Real deferred tax logic is complex.
        # Let's improve: If Deferred, TP = Prev_TP + Tax_Exp.
        if i == 0:
             ws.cell(row=row_idx, column=i+2, value=f"=IF({refs['tax_timing']}=0, 0, {col_letter}{tax_row})").number_format = currency_fmt
        else:
             prev_col = get_column_letter(i+1)
             ws.cell(row=row_idx, column=i+2, value=f"=IF({refs['tax_timing']}=0, 0, {prev_col}{row_idx}+{col_letter}{tax_row})").number_format = currency_fmt
             
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Long Term Debt")
    debt_row = row_idx
    # Update Interest Expense cells
    for cell in int_exp_cells:
        if "{DEBT_ROW}" in str(cell.value):
            cell.value = cell.value.replace("{DEBT_ROW}", str(debt_row))
            
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
             ws.cell(row=row_idx, column=i+2, value=f"={refs['debt']}").number_format = currency_fmt
        else:
             prev_col = get_column_letter(i+1)
             ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx}-{col_letter}{{REPAY_ROW}}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Common Stock")
    cs_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={refs['beg_cash']}+{refs['equity']}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Retained Earnings")
    re_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ni_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{re_row}+{col_letter}{ni_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Total Liab & Equity").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{ap_row}:{col_letter}{re_row})").number_format = currency_fmt
    row_idx += 3
    
    # --- CASH FLOW ---
    ws.cell(row=row_idx, column=1, value="CASH FLOW STATEMENT").font = bold_font
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Cash from Operations").font = bold_font
    row_idx += 1
    cfo_start_row = row_idx
    
    ws.cell(row=row_idx, column=1, value="Net Income")
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ni_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Depreciation")
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{deprec_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Change in AR")
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"=-{col_letter}{ar_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{ar_row}-{col_letter}{ar_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Change in Inventory")
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"=-{col_letter}{inv_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{inv_row}-{col_letter}{inv_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Change in AP")
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ap_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{ap_row}-{prev_col}{ap_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Change in Deferred Rev")
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{dr_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{dr_row}-{prev_col}{dr_row}").number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Change in Tax Payable")
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{tp_row}").number_format = currency_fmt
        else:
            prev_col = get_column_letter(i+1)
            ws.cell(row=row_idx, column=i+2, value=f"={col_letter}{tp_row}-{prev_col}{tp_row}").number_format = currency_fmt
    row_idx += 1
    cfo_end_row = row_idx - 1
    
    ws.cell(row=row_idx, column=1, value="Cash from Investing").font = bold_font
    row_idx += 1
    cfi_start_row = row_idx
    
    ws.cell(row=row_idx, column=1, value="CapEx")
    
    # Update Fixed Assets formulas
    for cell in fa_cells:
        cell.value = cell.value.replace("{CAPEX_ROW}", str(row_idx))
    
    # Formula: Sum of Cost Assumptions.
    capex_formula_parts = []
    for item in st.session_state.capex_items:
        capex_formula_parts.append(refs['capex'][item['name']]['cost'])
    capex_formula = "(" + "+".join(capex_formula_parts) + ")" if capex_formula_parts else "0"
    
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            formula = f"=-({capex_formula}+{col_letter}{total_rev_row}*{refs['maint_capex']})"
            ws.cell(row=row_idx, column=i+2, value=formula).number_format = currency_fmt
        else:
            formula = f"=-{col_letter}{total_rev_row}*{refs['maint_capex']}"
            ws.cell(row=row_idx, column=i+2, value=formula).number_format = currency_fmt
    row_idx += 1
    cfi_end_row = row_idx - 1
    
    ws.cell(row=row_idx, column=1, value="Cash from Financing").font = bold_font
    row_idx += 1
    cff_start_row = row_idx
    
    ws.cell(row=row_idx, column=1, value="Issuance of Common Stock")
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
             ws.cell(row=row_idx, column=i+2, value=f"={refs['beg_cash']}+{refs['equity']}").number_format = currency_fmt
        else:
             ws.cell(row=row_idx, column=i+2, value=0).number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Issuance of Debt")
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
             ws.cell(row=row_idx, column=i+2, value=f"={refs['debt']}").number_format = currency_fmt
        else:
             ws.cell(row=row_idx, column=i+2, value=0).number_format = currency_fmt
    row_idx += 1
    
    ws.cell(row=row_idx, column=1, value="Debt Repayment")
    repay_row = row_idx
    
    # Update Debt Row with Repayment Row
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        cell = ws.cell(row=debt_row, column=i+2)
        if "{REPAY_ROW}" in str(cell.value):
            cell.value = cell.value.replace("{REPAY_ROW}", str(repay_row))
            
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
             ws.cell(row=row_idx, column=i+2, value=0).number_format = currency_fmt
        else:
             prev_col = get_column_letter(i+1)
             # Repayment = Prev_Debt / (Term * 12) if Term > 0
             ws.cell(row=row_idx, column=i+2, value=f"=IF({refs['debt_term']}>0, {refs['debt']}/({refs['debt_term']}*12), 0)").number_format = currency_fmt
    row_idx += 1
    
    cff_end_row = row_idx - 1
    
    ws.cell(row=row_idx, column=1, value="Net Cash Flow").font = bold_font
    ncf_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=row_idx, column=i+2, value=f"=SUM({col_letter}{cfo_start_row}:{col_letter}{cfo_end_row})+SUM({col_letter}{cfi_start_row}:{col_letter}{cfi_end_row})+SUM({col_letter}{cff_start_row}:{col_letter}{cff_end_row})").number_format = currency_fmt
    row_idx += 2
    
    ws.cell(row=row_idx, column=1, value="Ending Cash Balance").font = bold_font
    ec_row = row_idx
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
             ws.cell(row=row_idx, column=i+2, value=f"=0+{col_letter}{ncf_row}").number_format = currency_fmt
        else:
             prev_col = get_column_letter(i+1)
             ws.cell(row=row_idx, column=i+2, value=f"={prev_col}{row_idx}+{col_letter}{ncf_row}").number_format = currency_fmt
             
    # Link BS Cash
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws.cell(row=cash_row, column=i+2, value=f"={col_letter}{ec_row}").number_format = currency_fmt
    
    ws.column_dimensions['A'].width = 30
    for col in ['B','C','D','E']:
        ws.column_dimensions[col].width = 15
        
    # 3. Annual Summary Sheet
    ws_summ = wb.create_sheet("Annual Summary")
    summ_headers = ["Item", "Year 1", "Year 2", "Year 3"]
    ws_summ.append(summ_headers)
    for cell in ws_summ[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        
    # Helper to sum 12 columns
    def add_summ_row(label, src_row, is_sum=True):
        r = ws_summ.max_row + 1
        ws_summ.cell(row=r, column=1, value=label)
        # Year 1: Sum B-M (2-13)
        # Year 2: Sum N-Y (14-25)
        # Year 3: Sum Z-AK (26-37)
        ranges = [('B', 'M'), ('N', 'Y'), ('Z', 'AK')]
        for i, (start, end) in enumerate(ranges):
            if is_sum:
                ws_summ.cell(row=r, column=i+2, value=f"=SUM('36 Month Model'!{start}{src_row}:{end}{src_row})").number_format = currency_fmt
            else:
                # For Balance Sheet items, take the ending value (last month of year)
                end_col = end
                ws_summ.cell(row=r, column=i+2, value=f"='36 Month Model'!{end_col}{src_row}").number_format = currency_fmt

    ws_summ.append(["PROFIT & LOSS"])
    ws_summ.cell(row=ws_summ.max_row, column=1).font = bold_font
    add_summ_row("Total Revenue", total_rev_row)
    add_summ_row("Total COGS", total_cogs_row)
    add_summ_row("Gross Profit", gross_profit_row)
    add_summ_row("Total OpEx", total_opex_row)
    add_summ_row("EBITDA", ebitda_row)
    add_summ_row("Depreciation", deprec_row)
    add_summ_row("EBIT", ebit_row)
    add_summ_row("Interest Expense", int_exp_row)
    add_summ_row("Interest Income", int_inc_row)
    add_summ_row("EBT", ebt_row)
    add_summ_row("Income Tax", tax_row)
    add_summ_row("Net Income", ni_row)
    
    ws_summ.append([""])
    ws_summ.append(["BALANCE SHEET"])
    ws_summ.cell(row=ws_summ.max_row, column=1).font = bold_font
    add_summ_row("Cash", cash_row, is_sum=False)
    add_summ_row("Accounts Receivable", ar_row, is_sum=False)
    add_summ_row("Inventory", inv_row, is_sum=False)
    add_summ_row("Total Assets", ad_row+1, is_sum=False) # Total Assets Row
    add_summ_row("Accounts Payable", ap_row, is_sum=False)
    add_summ_row("Long Term Debt", debt_row, is_sum=False)
    add_summ_row("Total Equity", re_row+1, is_sum=False) # Total Liab & Equity Row (approx)


    # 4. KPIs Sheet
    ws_kpi = wb.create_sheet("KPIs")
    
    headers_kpi = ["Metric"] + periods
    ws_kpi.append(headers_kpi)
    for cell in ws_kpi[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    kpi_row = 2
    
    # KPI Assumptions Export
    add_assump("KPIs", "Starting Customers", st.session_state.kpi_assumptions['starting_customers'][scen], None, 'start_cust')
    add_assump("KPIs", "New Customers Monthly", st.session_state.kpi_assumptions['new_customers_monthly'][scen], None, 'new_cust')
    add_assump("KPIs", "Monthly Churn Rate", st.session_state.kpi_assumptions['churn_rate_monthly'][scen], pct_fmt, 'churn_rate')
    
    # Customer Count
    ws_kpi.cell(row=kpi_row, column=1, value="Customer Count").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i == 0:
            ws_kpi.cell(row=kpi_row, column=i+2, value=f"={refs['start_cust']}+{refs['new_cust']}-(({refs['start_cust']})*{refs['churn_rate']})").number_format = '#,##0'
        else:
            prev_col = get_column_letter(i+1)
            ws_kpi.cell(row=kpi_row, column=i+2, value=f"={prev_col}{kpi_row}+{refs['new_cust']}-(({prev_col}{kpi_row})*{refs['churn_rate']})").number_format = '#,##0'
    cust_count_row = kpi_row
    kpi_row += 1
    
    # MRR (Monthly Recurring Revenue)
    ws_kpi.cell(row=kpi_row, column=1, value="MRR").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws_kpi.cell(row=kpi_row, column=i+2, value=f"='36 Month Model'!{col_letter}{total_rev_row}").number_format = currency_fmt
    mrr_row = kpi_row
    kpi_row += 1
    
    # S&M Spend (sum of OpEx items marked as S&M)
    ws_kpi.cell(row=kpi_row, column=1, value="S&M Spend").font = bold_font
    sm_opex_list = st.session_state.kpi_assumptions.get('sm_opex_items', [])
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        # Create formula to sum S&M opex items from 36 Month Model
        sm_refs = []
        for opex_idx, item in enumerate(st.session_state.opex_items):
            if item['name'] in sm_opex_list:
                sm_row = opex_start_row + opex_idx
                sm_refs.append(f"'36 Month Model'!{col_letter}{sm_row}")
        if sm_refs:
            formula = "=" + "+".join(sm_refs)
        else:
            formula = "=0"
        ws_kpi.cell(row=kpi_row, column=i+2, value=formula).number_format = currency_fmt
    sm_spend_row = kpi_row
    kpi_row += 1
    
    # CAC (Customer Acquisition Cost)
    ws_kpi.cell(row=kpi_row, column=1, value="CAC").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws_kpi.cell(row=kpi_row, column=i+2, value=f"=IF({refs['new_cust']}>0, {col_letter}{sm_spend_row}/{refs['new_cust']}, 0)").number_format = currency_fmt
    cac_row = kpi_row
    kpi_row += 1
    
    # Gross Margin %
    ws_kpi.cell(row=kpi_row, column=1, value="Gross Margin %").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws_kpi.cell(row=kpi_row, column=i+2, value=f"=IF('36 Month Model'!{col_letter}{total_rev_row}>0, '36 Month Model'!{col_letter}{gross_profit_row}/'36 Month Model'!{col_letter}{total_rev_row}, 0)").number_format = pct_fmt
    gm_pct_row = kpi_row
    kpi_row += 1
    
    # ARPA (Average Revenue Per Account)
    ws_kpi.cell(row=kpi_row, column=1, value="ARPA").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws_kpi.cell(row=kpi_row, column=i+2, value=f"=IF({col_letter}{cust_count_row}>0, {col_letter}{mrr_row}/{col_letter}{cust_count_row}, 0)").number_format = currency_fmt
    arpa_row = kpi_row
    kpi_row += 1
    
    # LTV (Lifetime Value)
    ws_kpi.cell(row=kpi_row, column=1, value="LTV").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws_kpi.cell(row=kpi_row, column=i+2, value=f"=IF({refs['churn_rate']}>0, ({col_letter}{arpa_row}*{col_letter}{gm_pct_row})/{refs['churn_rate']}, 0)").number_format = currency_fmt
    ltv_row = kpi_row
    kpi_row += 1
    
    # LTV:CAC Ratio
    ws_kpi.cell(row=kpi_row, column=1, value="LTV:CAC Ratio").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws_kpi.cell(row=kpi_row, column=i+2, value=f"=IF({col_letter}{cac_row}>0, {col_letter}{ltv_row}/{col_letter}{cac_row}, 0)").number_format = '0.00'
    ltv_cac_row = kpi_row
    kpi_row += 1
    
    # Gross Profit per Customer
    ws_kpi.cell(row=kpi_row, column=1, value="Gross Profit per Customer").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws_kpi.cell(row=kpi_row, column=i+2, value=f"=IF({col_letter}{cust_count_row}>0, '36 Month Model'!{col_letter}{gross_profit_row}/{col_letter}{cust_count_row}, 0)").number_format = currency_fmt
    gp_per_cust_row = kpi_row
    kpi_row += 1
    
    # CAC Payback (Months to Recover CAC)
    ws_kpi.cell(row=kpi_row, column=1, value="CAC Payback (Months)").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws_kpi.cell(row=kpi_row, column=i+2, value=f"=IF({col_letter}{gp_per_cust_row}>0, {col_letter}{cac_row}/{col_letter}{gp_per_cust_row}, 0)").number_format = '0.0'
    cac_payback_row = kpi_row
    kpi_row += 1
    
    # Revenue Growth % (YoY)
    ws_kpi.cell(row=kpi_row, column=1, value="Revenue Growth % (YoY)").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i < 12:
            ws_kpi.cell(row=kpi_row, column=i+2, value="N/A")
        else:
            prev_year_col = get_column_letter(i+2-12)
            ws_kpi.cell(row=kpi_row, column=i+2, value=f"=IF('36 Month Model'!{prev_year_col}{total_rev_row}>0, ('36 Month Model'!{col_letter}{total_rev_row}-'36 Month Model'!{prev_year_col}{total_rev_row})/'36 Month Model'!{prev_year_col}{total_rev_row}, 0)").number_format = pct_fmt
    rev_growth_row = kpi_row
    kpi_row += 1
    
    # EBITDA Margin %
    ws_kpi.cell(row=kpi_row, column=1, value="EBITDA Margin %").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        ws_kpi.cell(row=kpi_row, column=i+2, value=f"=IF('36 Month Model'!{col_letter}{total_rev_row}>0, '36 Month Model'!{col_letter}{ebitda_row}/'36 Month Model'!{col_letter}{total_rev_row}, 0)").number_format = pct_fmt
    ebitda_margin_row = kpi_row
    kpi_row += 1
    
    # Rule of 40
    ws_kpi.cell(row=kpi_row, column=1, value="Rule of 40").font = bold_font
    for i, p in enumerate(periods):
        col_letter = get_column_letter(i+2)
        if i < 12:
            ws_kpi.cell(row=kpi_row, column=i+2, value="N/A")
        else:
            ws_kpi.cell(row=kpi_row, column=i+2, value=f"={col_letter}{rev_growth_row}+{col_letter}{ebitda_margin_row}").number_format = pct_fmt
    rule40_row = kpi_row
    
    # Column sizing for KPIs will be done later with other sheets

    # --- SMART COLUMN SIZING ---
    
    # 1. Assumptions Sheet (Static Values)
    # Estimate width based on formatted value
    for col in ws_assump.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = cell.value
                fmt = cell.number_format
                if val is not None:
                    if isinstance(val, (int, float)):
                        if fmt == currency_fmt: # '#,##0.00'
                            l = len("{:,.2f}".format(val))
                        elif fmt == pct_fmt: # '0.00%'
                            l = len("{:.2%}".format(val))
                        else:
                            l = len(str(val))
                    else:
                        l = len(str(val))
                    max_len = max(max_len, l)
            except: pass
        ws_assump.column_dimensions[col_letter].width = max_len + 2

    # 2. Main & Summary Sheets (Formulas)
    # Calculate required width based on maximum financial assumptions
    max_val = 0.0
    
    # Check Revenue (Year 1 + Growth projection)
    for item in st.session_state.revenue_items:
        v = item['value'][scen]
        # Use the highest growth rate for width calculation
        g = max(
            item.get('growth_y1', {}).get(scen, 0.10),
            item.get('growth_y2', {}).get(scen, 0.07),
            item.get('growth_y3', {}).get(scen, 0.04)
        )
        # Project to Year 3 (approx)
        v_y3 = v * ((1+g)**3)
        max_val = max(max_val, v_y3)
        
    # Check Financing
    max_val = max(max_val, st.session_state.financing_assumptions['equity_raised'][scen])
    max_val = max(max_val, st.session_state.financing_assumptions['debt_issued'][scen])
    
    # Safety buffer for totals (e.g. Total Revenue > Single Stream)
    # 5x buffer covers most aggregations
    safe_max_val = max_val * 5 if max_val > 0 else 1000000 
    
    # Calculate width needed for this max value
    data_width = len("{:,.2f}".format(safe_max_val)) + 3 # +3 padding for safety
    
    # Apply to Main Sheet
    ws.column_dimensions['A'].width = 30 # Label column
    for i in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = data_width
        
    # Apply to Summary Sheet
    ws_summ.column_dimensions['A'].width = 30
    for i in range(2, ws_summ.max_column + 1):
        ws_summ.column_dimensions[get_column_letter(i)].width = data_width
    
    # Apply to KPIs Sheet
    ws_kpi.column_dimensions['A'].width = 30
    for i in range(2, ws_kpi.max_column + 1):
        ws_kpi.column_dimensions[get_column_letter(i)].width = data_width

    return wb

# --- TITLE & CREDITS ---
st.title("Dynamic 3-Statement Financial Model")
st.markdown("Made by [Avishek Kumar Jaiswal](https://www.linkedin.com/in/avishek-kumar-jaiswal/)")
st.markdown("---")

# --- DOWNLOAD BUTTON ---
try:
    wb = generate_excel()
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    col_dl1, col_dl2 = st.columns([3, 1])
    with col_dl2:
        st.download_button(
            label="📥 Download Dynamic Model (.xlsx)",
            data=buffer,
            file_name="Dynamic_Financial_Model.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
except Exception as e:
    st.error(f"Error generating Excel file: {e}")
    
    
col_scen1, col_scen2 = st.columns(2)
with col_scen1:
    st.session_state.scenario_to_edit = st.selectbox("Scenario to Edit", SCENARIOS, index=SCENARIOS.index(st.session_state.scenario_to_edit))
with col_scen2:
    st.session_state.scenario_to_run = st.selectbox("Scenario to Run (Excel)", SCENARIOS, index=SCENARIOS.index(st.session_state.scenario_to_run))

st.info(f"Editing values for: **{st.session_state.scenario_to_edit}**")
curr_scen = st.session_state.scenario_to_edit

col_main1, col_main2 = st.columns(2)

with col_main1:
    # 1. Revenue Streams
    with st.expander("1. Revenue Streams", expanded=True):
        for i, item in enumerate(st.session_state.revenue_items):
            st.markdown(f"**{item['name']}**")
            c1, c2 = st.columns(2)
            item['value'][curr_scen] = c1.number_input(f"Year 1 Value ($) ##{i}", value=item['value'][curr_scen], step=1000.0, key=f"rev_val_{i}_{curr_scen}")
            # Growth Tapering: Separate rates for each year
            st.markdown("**Growth Rates (Annual)**")
            c1, c2, c3 = st.columns(3)
            item['growth_y1'] = item.get('growth_y1', init_scenario_val(0.10))
            item['growth_y2'] = item.get('growth_y2', init_scenario_val(0.07))
            item['growth_y3'] = item.get('growth_y3', init_scenario_val(0.04))
            item['growth_y1'][curr_scen] = c1.number_input(f"Year 1 (%) ##{i}", value=item['growth_y1'][curr_scen]*100, step=1.0, key=f"rev_gr_y1_{i}_{curr_scen}") / 100
            item['growth_y2'][curr_scen] = c2.number_input(f"Year 2 (%) ##{i}", value=item['growth_y2'][curr_scen]*100, step=1.0, key=f"rev_gr_y2_{i}_{curr_scen}") / 100
            item['growth_y3'][curr_scen] = c3.number_input(f"Year 3 (%) ##{i}", value=item['growth_y3'][curr_scen]*100, step=1.0, key=f"rev_gr_y3_{i}_{curr_scen}") / 100
            if st.button(f"Remove {item['name']}", key=f"del_rev_{i}"):
                st.session_state.revenue_items.pop(i)
                st.rerun()
        
        st.markdown("---")
        new_rev_name = st.text_input("New Revenue Name", key="new_rev_name")
        if st.button("Add Revenue Stream"):
            if new_rev_name:
                st.session_state.revenue_items.append({
                    'name': new_rev_name, 
                    'value': init_scenario_val(0.0), 
                    'growth_y1': init_scenario_val(0.10),
                    'growth_y2': init_scenario_val(0.07),
                    'growth_y3': init_scenario_val(0.04)
                })
                st.rerun()

    # 3. OpEx
    with st.expander("3. OpEx (People vs Fixed)", expanded=False):
        for i, item in enumerate(st.session_state.opex_items):
            st.markdown(f"**{item['name']}**")
            type_opts = ["Fixed Amount", "% of Rev", "Personnel"]
            curr_type_idx = type_opts.index(item.get('type', 'Fixed Amount'))
            item['type'] = st.selectbox(f"Type ##{i}", type_opts, index=curr_type_idx, key=f"opex_type_{i}")
            
            c1, c2 = st.columns(2)
            if item['type'] == "Fixed Amount":
                item['value'][curr_scen] = c1.number_input(f"Fixed Amount ($) ##{i}", value=float(item['value'][curr_scen]), step=1000.0, key=f"opex_val_{i}_{curr_scen}")
                item['param2'][curr_scen] = c2.number_input(f"Growth (%) ##{i}", value=float(item.get('param2', {}).get(curr_scen, 0.0))*100, step=1.0, key=f"opex_p2_{i}_{curr_scen}") / 100
            elif item['type'] == "% of Rev":
                item['value'][curr_scen] = c1.number_input(f"% of Revenue ##{i}", value=float(item['value'][curr_scen])*100, step=1.0, key=f"opex_val_{i}_{curr_scen}") / 100
                # param2 unused for % of Rev
            elif item['type'] == "Personnel":
                item['value'][curr_scen] = c1.number_input(f"Headcount (Start) ##{i}", value=float(item['value'][curr_scen]), step=1.0, key=f"opex_val_{i}_{curr_scen}")
                item['param2'][curr_scen] = c2.number_input(f"Avg Salary ($/yr) ##{i}", value=float(item.get('param2', {}).get(curr_scen, 50000.0)), step=1000.0, key=f"opex_p2_{i}_{curr_scen}")
                # Dynamic Headcount: Revenue threshold for hiring
                st.markdown("**Dynamic Headcount Settings**")
                item['revenue_threshold'] = item.get('revenue_threshold', init_scenario_val(50000.0))
                item['revenue_threshold'][curr_scen] = st.number_input(
                    f"Revenue per New Hire ($) ##{i}", 
                    value=float(item.get('revenue_threshold', {}).get(curr_scen, 50000.0)), 
                    step=5000.0, 
                    key=f"opex_threshold_{i}_{curr_scen}",
                    help="Hire 1 additional person for every $X increase in monthly revenue"
                )

            if st.button(f"Remove {item['name']}", key=f"del_opex_{i}"):
                st.session_state.opex_items.pop(i)
                st.rerun()

        st.markdown("---")
        new_opex_name = st.text_input("New OpEx Name", key="new_opex_name")
        if st.button("Add OpEx Item"):
            if new_opex_name:
                st.session_state.opex_items.append({'name': new_opex_name, 'value': init_scenario_val(5000.0), 'type': 'Fixed Amount', 'param2': init_scenario_val(0.0)})
                st.rerun()

    # 5. Assets & Taxation
    with st.expander("5. Assets & Taxation", expanded=False):
        st.markdown("**Assets**")
        for i, item in enumerate(st.session_state.capex_items):
            st.markdown(f"**{item['name']}**")
            c1, c2 = st.columns(2)
            item['cost'][curr_scen] = c1.number_input(f"Cost ($) ##{i}", value=item['cost'][curr_scen], step=500.0, key=f"capex_cost_{i}_{curr_scen}")
            item['deprec_rate'][curr_scen] = c2.number_input(f"Deprec Rate (%) ##{i}", value=item.get('deprec_rate', {}).get(curr_scen, 0.20)*100, step=1.0, key=f"capex_rate_{i}_{curr_scen}") / 100
            if st.button(f"Remove {item['name']}", key=f"del_capex_{i}"):
                st.session_state.capex_items.pop(i)
                st.rerun()
        
        new_capex_name = st.text_input("New Asset Name", key="new_capex_name")
        if st.button("Add Asset"):
            if new_capex_name:
                st.session_state.capex_items.append({'name': new_capex_name, 'cost': init_scenario_val(1000.0), 'deprec_rate': init_scenario_val(0.20)})
                st.rerun()
        
        st.markdown("---")
        st.markdown("**Maintenance CapEx**")
        st.session_state.capex_assumptions['maintenance_pct'][curr_scen] = st.slider(
            "Ongoing CapEx (% of Monthly Revenue)", 
            0.0, 10.0, 
            float(st.session_state.capex_assumptions.get('maintenance_pct', {}).get(curr_scen, 2.0)*100), 
            0.1,
            key=f"maint_capex_{curr_scen}",
            help="Monthly capital expenditure as a percentage of revenue"
        ) / 100
        
        st.markdown("---")
        st.markdown("**Taxation**")
        st.session_state.tax_assumptions['tax_rate'][curr_scen] = st.slider("Tax Rate (%)", 0, 50, int(st.session_state.tax_assumptions['tax_rate'][curr_scen]*100), key=f"tax_rate_{curr_scen}") / 100
        st.session_state.tax_assumptions['payment_timing'] = st.radio("Tax Payment Timing", ["Immediate", "Next Year"], index=0 if st.session_state.tax_assumptions['payment_timing'] == "Immediate" else 1)
        st.session_state.tax_assumptions['nol_balance'] = st.number_input("NOL Beginning Balance ($)", value=st.session_state.tax_assumptions.get('nol_balance', 0.0), step=1000.0)


with col_main2:
    # 2. COGS
    with st.expander("2. COGS", expanded=False):
        for i, item in enumerate(st.session_state.cogs_items):
            st.markdown(f"**{item['name']}**")
            c1, c2 = st.columns(2)
            item['type'] = c1.selectbox(f"Type ##{i}", ["% of Rev", "Fixed Amount"], index=0 if item['type'] == "% of Rev" else 1, key=f"cogs_type_{i}")
            
            if item['type'] == "% of Rev":
                val = c2.number_input(f"% of Rev ##{i}", value=item['value'][curr_scen]*100, step=1.0, key=f"cogs_val_{i}_{curr_scen}") / 100
            else:
                val = c2.number_input(f"Fixed ($) ##{i}", value=item['value'][curr_scen], step=500.0, key=f"cogs_val_{i}_{curr_scen}")
            item['value'][curr_scen] = val
            
            if st.button(f"Remove {item['name']}", key=f"del_cogs_{i}"):
                st.session_state.cogs_items.pop(i)
                st.rerun()

        st.markdown("---")
        new_cogs_name = st.text_input("New COGS Name", key="new_cogs_name")
        if st.button("Add COGS Item"):
            if new_cogs_name:
                st.session_state.cogs_items.append({'name': new_cogs_name, 'value': init_scenario_val(0.10), 'type': '% of Rev'})
                st.rerun()

    # 4. Working Capital
    with st.expander("4. Working Capital", expanded=False):
        st.session_state.wc_assumptions['beginning_cash'] = st.number_input("Beginning Cash Balance ($)", value=st.session_state.wc_assumptions['beginning_cash'], step=1000.0)
        st.session_state.wc_assumptions['ar_percent'][curr_scen] = st.slider("AR as % of Rev (DSO Proxy)", 0, 50, int(st.session_state.wc_assumptions['ar_percent'][curr_scen]*100), key=f"ar_{curr_scen}") / 100
        st.session_state.wc_assumptions['ap_percent'][curr_scen] = st.slider("AP as % of Opex (DPO Proxy)", 0, 50, int(st.session_state.wc_assumptions['ap_percent'][curr_scen]*100), key=f"ap_{curr_scen}") / 100
        st.session_state.wc_assumptions['deferred_rev_percent'][curr_scen] = st.slider("Deferred Revenue (% of Sales)", 0, 50, int(st.session_state.wc_assumptions.get('deferred_rev_percent', {}).get(curr_scen, 0.0)*100), key=f"dr_{curr_scen}") / 100
        st.markdown("**Inventory**")
        st.session_state.wc_assumptions['days_inventory'][curr_scen] = st.number_input("Days Inventory Outstanding (DIO)", value=st.session_state.wc_assumptions.get('days_inventory', {}).get(curr_scen, 30.0), step=1.0, key=f"dio_{curr_scen}")
        st.session_state.wc_assumptions['days_payable'][curr_scen] = st.number_input("Days Payable Outstanding (DPO)", value=st.session_state.wc_assumptions.get('days_payable', {}).get(curr_scen, 30.0), step=1.0, key=f"dpo_{curr_scen}")

    # 6. Financing
    with st.expander("6. Financing", expanded=False):
        st.session_state.financing_assumptions['equity_raised'][curr_scen] = st.number_input("New Equity Raised ($)", value=st.session_state.financing_assumptions['equity_raised'][curr_scen], step=1000.0, key=f"eq_{curr_scen}")
        st.session_state.financing_assumptions['debt_issued'][curr_scen] = st.number_input("New Debt Issued ($)", value=st.session_state.financing_assumptions['debt_issued'][curr_scen], step=1000.0, key=f"debt_{curr_scen}")
        st.session_state.financing_assumptions['debt_interest_rate'][curr_scen] = st.number_input("Interest Rate on Debt (%)", value=st.session_state.financing_assumptions['debt_interest_rate'][curr_scen]*100, step=0.1, key=f"d_int_{curr_scen}") / 100
        st.session_state.financing_assumptions['cash_interest_rate'][curr_scen] = st.number_input("Interest on Cash (%)", value=st.session_state.financing_assumptions['cash_interest_rate'][curr_scen]*100, step=0.1, key=f"c_int_{curr_scen}") / 100
        st.session_state.financing_assumptions['overdraft_interest_rate'][curr_scen] = st.number_input("Overdraft Interest Rate (%)", value=st.session_state.financing_assumptions.get('overdraft_interest_rate', {}).get(curr_scen, 0.10)*100, step=0.1, key=f"od_int_{curr_scen}") / 100
        st.session_state.financing_assumptions['debt_repayment_term'][curr_scen] = st.number_input("Debt Repayment Term (Years)", value=int(st.session_state.financing_assumptions.get('debt_repayment_term', {}).get(curr_scen, 5)), step=1, key=f"term_{curr_scen}")


    # 7. KPI Assumptions
    with st.expander("7. KPI Assumptions", expanded=False):
        st.markdown("**Customer Metrics**")
        st.session_state.kpi_assumptions['starting_customers'][curr_scen] = st.number_input(
            "Starting Customer Count", 
            value=float(st.session_state.kpi_assumptions['starting_customers'][curr_scen]), 
            step=10.0, 
            key=f"start_cust_{curr_scen}"
        )
        st.session_state.kpi_assumptions['new_customers_monthly'][curr_scen] = st.number_input(
            "New Customers per Month", 
            value=float(st.session_state.kpi_assumptions['new_customers_monthly'][curr_scen]), 
            step=5.0, 
            key=f"new_cust_{curr_scen}",
            help="Average number of new customers acquired each month"
        )
        st.session_state.kpi_assumptions['churn_rate_monthly'][curr_scen] = st.number_input(
            "Monthly Churn Rate (%)", 
            value=float(st.session_state.kpi_assumptions['churn_rate_monthly'][curr_scen])*100, 
            step=0.5, 
            key=f"churn_{curr_scen}",
            help="Percentage of customers lost each month"
        ) / 100
        
        st.markdown("---")
        st.markdown("**Sales & Marketing Classification**")
        st.info("Select which OpEx items should be counted as Sales & Marketing spend for CAC calculation")
        
        sm_items = st.session_state.kpi_assumptions.get('sm_opex_items', [])
        for item in st.session_state.opex_items:
            is_sm = st.checkbox(
                f"{item['name']} is Sales & Marketing", 
                value=item['name'] in sm_items,
                key=f"sm_{item['name']}_{curr_scen}"
            )
            if is_sm and item['name'] not in sm_items:
                sm_items.append(item['name'])
            elif not is_sm and item['name'] in sm_items:
                sm_items.remove(item['name'])
        st.session_state.kpi_assumptions['sm_opex_items'] = sm_items

st.markdown("---")